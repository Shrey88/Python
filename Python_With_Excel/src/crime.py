""" *
    *   one of the main drawback of openpyxl is that it is very slow when working with very large dataset
    *   because it carries with it all the tools that graphically enhances the worksheet even while parsing data
    *   in case like where we have to analyse the large datasets, its easy to use pandas.
    *
    * """
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook(".\\Exercise Files\\template.xlsx")
ws = wb.active

""" *
    *   next we will create a dataframe as it is reading a file so large, we need to add few flags 
    *   we will be adding encoding flag so that it knows in what encoding to read from
    *   some casting to data type of file
    *
    *   when handling the large files like crime.csv, datatype casting is always necessary, the reason for this is to
    *   conserve memory
    *
    *   when initially reading the file, panda tries to guess what type of object say integer or string that each column
    *   should cast to based on the values in it, but since we have specified datatypes already, pandas does not have to 
    *   worry about it when reading a file
    * """
df = pd.read_csv('.\\Exercise Files\\crime.csv', encoding='UTF-8', dtype={
    "INCIDENT_NUMBER":str, "OFFENSE_CODE":str, "OFFENSE_CODE_GROUP":str, "OFFENSE_DESCRIPTION":str,
    "DISTRICT":str, "REPORTING_AREA":str, "SHOOTING":str, "YEAR":str, "MONTH":str, "DAY_OF_WEEK":str,
    "HOUR":str
})


#   next step is to trim or filter a data
#   this will only be a fraction of size than the old one
df1 = df[df['OFFENSE_CODE_GROUP'] == 'Counterfeiting']


#   to clean this up some more we can use numpy to replace our name values
df1 = df1.replace(np.nan, "N/A", regex=True)


#   let's take total no of crimes and total no of counterfeit and then take the percentage
total_crimes = len(df.index)
total_counterfeit = len(df1.index)

#   to create percentage we will use python's built-in round function
perc_crimes = (total_counterfeit / total_crimes)*100
perc_crimes = round(perc_crimes, 2)

ws['O8'].value = total_crimes
ws['P8'].value = total_counterfeit
ws['Q8'].value = perc_crimes

#   count the no of crimes appearing each year in district
df1['Count'] = 1
df2 = df1.groupby(['DISTRICT', 'YEAR']).count()['Count']

#   year value repeats over and over for each district, we want something that makes it look more like a indexes.
print(df2)

#   we can add unstack function to make year look more like an indexes
df2 = df1.groupby(['DISTRICT', 'YEAR']).count()['Count'].unstack(level=0)
print(df2)

""" *
    *   only problem is N/A column because NaN value cannot be changed to str value N/A, we can simply use dataframe 
    *   drop function to get rid of the problem.
    * """
df2.drop(columns='N/A', inplace=True)

#   now we can convert the data in openpyxl format
rows = dataframe_to_rows(df2)
for r_idx, row in enumerate(rows, 8):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(".\\Exercise Files\\crime_report.xlsx")