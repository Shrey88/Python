""" *
    *   when working with pandas it's easy to create a column in dataframes
    * """
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
#   we will create 3 different dataframes for 3 different worksheets
df_1 = pd.read_excel(".\\Exercise Files\\shifts.xlsx", sheet_name="Sheet")
df_2 = pd.read_excel(".\\Exercise Files\\shifts.xlsx", sheet_name="Sheet1")
df_3 = pd.read_excel(".\\Exercise Files\\shift_3.xlsx", sheet_name="Sheet1")

#   we need to use pandas concate function to combine all the dataframes as long as all of them have same columns
#   provide the list of dataframes as the input and set the sort as False so that columns remains in same order
df_all = pd.concat([df_1, df_2, df_3], sort=False)
print(df_all)

#   let's print out row no 50
#   this will give you the data from row no 50 from all the shifts
print(df_all.loc[50])

#   you can use the group by function on column Shift to analyse the mean of Units sold
print(df_all.groupby(['Shift']).mean()['Units Sold'])

to_excel =df_all.to_excel(".\\Exercise Files\\all_shifts.xlsx", index=None)

""" *
    *   in the new excel we want to add new column Total representing the amount each sales representative has made
    *   
    * """
wb = load_workbook(".\\Exercise Files\\all_shifts.xlsx")
ws = wb.active

#   we will reference the cell that we want to use as header for our Total column
total_col = ws['G1']
total_col.font = Font(bold=True)
total_col.value = 'Total'

#   we know that the values of cost per is E and UnitsSold is in F, we will create a variable that stores the the column
#   name
e_col, f_col = ['E', 'F']


#   we will iterate through the total no of rows skipping the first row as it contains the header information
for row in range(2, 300):
    result_cell = 'G{}'.format(row)
    e_value = ws[e_col + str(row)].value
    f_value = ws[f_col + str(row)].value
    ws[result_cell] = e_value * f_value

#   save our workbook
wb.save('.\\Exercise Files\\totaled.xlsx')