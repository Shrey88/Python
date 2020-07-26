""" *
    *   we used pandas to open the file perform some analysis or statistics and then export to excel. However
    *   sometimes we cannot just use pandas built-in accelerator to overwrite a saved file or we want to place our data
    *   begins, this is where openpyxl comes in
    * """
import pandas as pd
from openpyxl import load_workbook
#   this will allow us to convert our dataframe rows into the format that is usable by openpyxl.
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook(".\\Exercise Files\\regions.xlsx")
ws = wb.active


#   now let us get our data into pandas
df = pd.read_excel('.\\Exercise Files\\all_shifts.xlsx')
df1 = df[['Sales Rep', 'Cost per', 'Units Sold']]
print(df1)


""" *
    *   last step is to convert dataframe with which openpyxl can work with so that we don't overwrite any cells
    *
    *   we are not going to make sure that we don't overwrite the existing data in the excel, so we need to use 
    *   openpyxl integration to convert from dataframe to rows
    *
    *   we are going to set the indexes as false as we don't need them
    * """
rows = dataframe_to_rows(df1, index=False)
print(rows)

for r_idx, row in enumerate(rows, 1):
    for c_idx, col in enumerate(row, 6):
        ws.cell(row=r_idx, column=c_idx, value=col)

wb.save(".\\Exercise Files\\combines.xlsx")