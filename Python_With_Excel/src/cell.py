""" *
    *
    * """
from openpyxl.workbook import workbook
from openpyxl import load_workbook

wb = load_workbook('.\\Exercise Files\\regions.xlsx')
ws = wb.active

""" *
    *   you can access the column and row range by providing the index position of starting cell sliced by index 
    *   position of cell until which we want the rows or column to be selected.
    *   
    *   cells can be accessed as string value for column and row
    * """
cell_range = ws['A1':'C1']  # <-- it creates a tuple of cell objects from A1 to C1 i.e A1, B1, C1
print(cell_range)
print('=' * 20)
print()

# you can also use cell method
# i.e. A1 = 1:1     B1 = 1:2    C1 = 1:3
cell_range = ws.cell(row=1, column=2).value
print(cell_range)
print('=' * 20)
print()

""" *
    *   but if you want to select a column you just need to provide the column name or int value as index
    * """
cell_range = ws['C']  # <-- you can think this as an single dimensional array
print(cell_range)
print('=' * 20)
print()

cell_range = ws['C4']  # <-- you can select a single cell
print(cell_range)
print('=' * 20)
print()

#  you can also provide the range of columns
cell_range = ws['A':'C']
print(cell_range)
print('=' * 20)
print()

#  so if you mention this way it will give you cell C5 because the row integer index starts from 0
# i.e. A1 = ['A'][0]     B1 = ['B'][0]    C1 = ['C'][0]
cell_range = ws['C'][4]
print(cell_range)
print('=' * 20)
print()

#  you can access the cell using index here the row index starts from 1
# i.e. A1 = 1:0     B1 = 1:1    C1 = 1:2
cell_range = ws[1][0]
print(cell_range)
print('=' * 20)
print()


""" *
    *   cells are accessed by combination of string header and integers and columns are only the headers.
    *   for rows we use only integers
    *"""
row_range = ws[1:5]  # <-- in this we are selecting all the columns and 5 rows
print(row_range)
print('=' * 20)
print()


""" *
    *   iterating through columns and rows, when iterating through this object we use specific function in our for
    *   loop for row (ws.iter_rows) and column (ws.iter_column) function
    * """

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
print('=' * 20)
print()

""" *
    *   we can also specify flag i.e value_only=True along with max_row,min_row, max_col this will print only the 
    *   values instead of object information
    * """

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)
print('=' * 20)
print()