from openpyxl.workbook import Workbook
from openpyxl import load_workbook


""" *
    *   create two most important objects
    *   1. workbook 
    *   2. worksheet
    *
    *   wb is for workbook
    *   ws is for worksheet
    * """
wb = Workbook()
ws = wb.active  # <-- this creates a worksheet reference equal to the active sheet in our current workbook


""" *
    *   we will create new worksheets over active sheet
    * """
ws1 = wb.create_sheet('New Sheet')
# all it does is count an index, so in order of our worksheets this one appears before everything else
ws2 = wb.create_sheet('Another', 0)


# altering the title of our active worksheet, we can do this using title function.
ws.title = 'MySheet'

# printing the all the sheet names from our workbook
print(wb.sheetnames)

""" *
    *   one cool tidbit is that openpyxl won't actually create cells on your worksheet as objects until they're 
    *   accessed. so creating and writing a new workbook like this doesn't take too much memory.
    *   
    *   let's load the existing workbook to access its cells.
    * """
wb2 = load_workbook(".\\Exercise Files\\regions.xlsx")

new_sheet = wb2.create_sheet('NewSheet')

""" *
    *   at this point we start accessing cells, in an Excel spreadsheet cell objects are indexed by their column 
    *   value which is typically A through Z and their row value which starts at one. Now if you remember our worksheet 
    *   is the object that contains all of the cell objects. So if we want to grab the cell in the top left, 
    *   we know it should be cell A1and we need to get it through our worksheet. 
    *   Openpyxl allows us to do access this by using the column and row name like an index.
    * """
active_sheet = wb2.active
cell = active_sheet['A1']  # <-- we get the object notation of cell
print(cell)
print(cell.value)  # <-- we are printing the value of cell.

# if we want to change the value of cell
active_sheet['A1'] = 0
wb2.save('.\\Exercise Files\\modified_excel.xlsx')