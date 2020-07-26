""" *
    *   Formatting Workbook
    * """
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


wb = Workbook()
ws = wb.active

""" *
    *   to show some styling options in our spreadsheet we are going to loop some values
    * """
for i in range(1,20):
    ws.append(range(300))

""" *
    *   you can merge and unmerge cells using string index
    * """
ws.merge_cells("A1:B5")
ws.unmerge_cells("A1:B5")
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

#  we can save this entire merged cell into a variable called cell by addressing the B2 position of the worksheet
cell = ws['B2']

#  changing the font color, size and font style
cell.font = Font(color=colors.RED, size=20, italic=True)
cell.value = 'Merged_Cell'

#  lets say we want the text on the bottom right corner of the cell, all we need to do is change the alignment
#  alignment object takes two parameters horizontal and vertical location
cell.alignment = Alignment(horizontal='right', vertical='bottom')


""" *
    *   let's make the background of the cell a nice smooth transition from black to white, to accomplish this we will 
    *   use the cell's fill property.
    *   we have to use one of the two fill functions we imported, if we want a solid color fill we could use cell's 
    *   pattern fill, but in order to achieve a nice transition let's use gradient fill.
    *   in the gradient fill we need to specify the colour to start at and colour to stop at. 
    
    *   In Openpyxl and excel documents colors are encoded in hexadecimal, in rrggbb format.
    *   this means that first two color corresponds to red, the second two to green and the third two to blue   
    * """
cell.fill = GradientFill(stop=("000000", "FFFFFF"))
wb.save(".\\Exercise Files\\text.xlsx")


""" *
    *   Named styles are objects we can create that store a style so that we can use it multiple times instead of having
    *   to type all that out
    *   Now this is initialized
    * """
highlight = NamedStyle(name='highlight')

# changing the properties
highlight.font = Font(bold=True)

""" *
    *   we will create the thicker borders for that we need to create a variable to hold the style of a side.
    *   We'll make this slide have a fixed style and be black in color, Once this is done we need to change our 
    *   named style's border property. We do this by setting each edge equal to our new style.
    * """
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

#   we will fill the cell with the solid yellow color using pattern fill
highlight.fill = PatternFill('solid', fgColor="FFFF00")

#   we want to apply this pattern fill diagonally starting at H column.
#   remember that our iter_col function column is stored as a tuple.
count = 0
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = highlight
    count += 1
wb.save(".\\Exercise Files\\hihglight.xlsx")