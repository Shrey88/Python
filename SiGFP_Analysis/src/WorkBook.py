from datetime import datetime,date
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet
from openpyxl.styles import Font, NumberFormatDescriptor
from openpyxl import chart


class WorkBook(object):

    def __init__(self, data: list = None, column_header: list = None, filename = None, output_file=None):
        if filename is None:
            self._wb = Workbook()
            self._ws = self._wb.active
        else:
            self._wb = load_workbook(filename)
            self._ws = self._wb.active

        self._data = data
        self._col_header = column_header
        self._output_file = output_file

    @property
    def wb_object(self):
        return self._wb

    @property
    def active_sheet(self):
        pass

    @active_sheet.getter
    def active_sheet(self):
        return self._ws

    @active_sheet.setter
    def active_sheet(self, ws: worksheet):
        if self._wb.active == ws:
            print("Worksheet is already active....")
        else:
            self._wb.active = ws

    def create_worksheet(self, sheet_name=None):
        if sheet_name is None:
            return self._wb.create_sheet()
        else:
            return self._wb.create_sheet(sheet_name)

    def add_chart(self, achart: chart):
        pass

    def __write_data(self):
        if self._data.__sizeof__() > 0:
            row_index = 2
            self._ws.append(self._col_header)
            for col in range(self._col_header.__len__()):
                self._first_row = self._ws.cell(1, col+1)
                self._first_row.font = Font(bold=True)

            for row in self._data:
                col_index = 0
                for attr in row:
                    if isinstance(attr, date):
                        cell: Cell = self._ws[row_index][col_index]
                        cell.value = attr
                    else:
                        cell: Cell = self._ws[row_index][col_index]
                        if col_index in range(3,6):
                            cell.value = int(attr)
                        else:
                            cell.value = attr

                    col_index += 1

                row_index += 1

        self._wb.save(self._output_file)

    def write_data(self):
        self.__write_data()